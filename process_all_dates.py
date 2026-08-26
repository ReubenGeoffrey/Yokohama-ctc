"""
ATC Tires - Minimalist Professional Yellow CTC Reconciliation & Attendance Mapper
Single Row (Unmerged) Employee Records - Columns A to I (NO Gap, Includes Total WOP Count)
"""

import os
import re
import calendar
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

MONTH_MAP = {'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,'JUL':7,'AUG':8,'SEP':9,'OCT':10,'NOV':11,'DEC':12}
FONT_NAME = 'Segoe UI'

# Minimalist Yellow Theme Colors
C_YELLOW_MAIN = 'FFE699'   # Soft warm professional yellow
C_YELLOW_SUB = 'FFF2CC'    # Pale pastel yellow
C_YELLOW_ACCENT = 'FDE68A' # Warm gold
C_YELLOW_LIGHT = 'FFFBEB'  # Tinted yellow
C_YELLOW_PALE = 'FEFCE8'   # Very pale cream
C_YELLOW_GOLD = 'FDE047'   # Gold highlight
C_YELLOW_TOTAL = 'FACC15'  # Deep amber gold
C_TEXT_DARK = '1F2937'     # Dark charcoal slate
C_TEXT_SUB = '374151'      # Medium slate
C_GRAY_HEADER = 'F3F4F6'
C_BORDER = 'D1D5DB'        # Subtle clean border

thin_side = Side(style='thin', color=C_BORDER)
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

def time_str_to_hours(v):
    if not v: return 0.0
    if isinstance(v, (int, float)): return float(v) * 24.0
    m = re.match(r'^(\d{1,2}):(\d{2})', str(v).strip())
    if m: return int(m.group(1)) + int(m.group(2))/60.0
    return 0.0

def extract_file_info(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws_name = 'Present' if 'Present' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[ws_name]
    rows = list(ws.iter_rows(values_only=True))
    
    file_date = None
    category = None
    h_idx = -1
    
    for i, r in enumerate(rows[:10]):
        if not r: continue
        r0 = str(r[0] or '').strip()
        
        for c in r[:8]:
            c_str = str(c or '').strip()
            if 'PRINTED ON' in c_str.upper() or 'PRINTED AT' in c_str.upper():
                continue
                
            m = re.search(r'as of\s+(\d{1,2})\w{0,2}\s+([A-Za-z]{3,})\s+(\d{4})', c_str, re.IGNORECASE)
            if m and not file_date:
                d = int(m.group(1))
                mon = MONTH_MAP.get(m.group(2)[:3].upper())
                yr = int(m.group(3))
                if mon: file_date = datetime(yr, mon, d)
                
            m2 = re.search(r'(\d{1,2})[-\s/]([A-Za-z]{3,})[-\s/](\d{4})', c_str)
            if m2 and not file_date:
                d = int(m2.group(1))
                mon = MONTH_MAP.get(m2.group(2)[:3].upper())
                yr = int(m2.group(3))
                if mon: file_date = datetime(yr, mon, d)

        cat_str = ' '.join([str(c or '').upper() for c in r[:8]])
        if 'CATEGORY :' in cat_str or 'CATEGORY:' in cat_str:
            if 'OPERATOR' in cat_str: category = 'OP'
            elif 'NAPS' in cat_str: category = 'NAPS'
            elif 'CONTRACT' in cat_str or 'CL' in cat_str: category = 'CL'
        if r0.upper() in ['SR.NO', 'SRNO', 'SNO']:
            h_idx = i

    if not file_date:
        fn = os.path.basename(filepath)
        m_dmy = re.search(r'(\d{1,2})\s*[-_/\.]\s*(\d{1,2})\s*[-_/\.]\s*(\d{4})', fn)
        if m_dmy:
            file_date = datetime(int(m_dmy.group(3)), int(m_dmy.group(2)), int(m_dmy.group(1)))
        else:
            m_iso = re.search(r'(\d{4})-(\d{2})-(\d{2})', fn)
            if m_iso:
                file_date = datetime(int(m_iso.group(1)), int(m_iso.group(2)), int(m_iso.group(3)))
            else:
                m_fn = re.search(r'Date\s*(\d{1,2})', fn, re.IGNORECASE) or re.search(r'(\d{1,2})[-_]Aug', fn, re.IGNORECASE)
                if m_fn:
                    file_date = datetime(2026, 8, int(m_fn.group(1)))

    if not category and h_idx != -1:
        h = [str(c or '').strip().upper() for c in rows[h_idx]]
        c_idx = h.index('CODE') if 'CODE' in h else 1
        for r in rows[h_idx+1:h_idx+6]:
            if r and len(r) > c_idx and r[c_idx]:
                code = str(r[c_idx]).strip().upper()
                if code.startswith('9'): category = 'OP'; break
                elif code.startswith('LN'): category = 'NAPS'; break
                else: category = 'CL'; break
                
    if not category:
        fn = os.path.basename(filepath).upper()
        if 'OPERATOR' in fn or 'OP' in fn: category = 'OP'
        elif 'NAPS' in fn: category = 'NAPS'
        else: category = 'CL'
        
    return file_date, category, rows, h_idx

def style_summary_sheet(ws, year, month):
    ws.freeze_panes = 'B4'
    ws.views.sheetView[0].showGridLines = True
    
    ws.row_dimensions[1].height = 25
    ws.merge_cells('B1:Q1')
    c_dir = ws['B1']
    c_dir.value = 'DIRECT LABOUR / PRODUCTION'
    c_dir.fill = PatternFill('solid', fgColor=C_YELLOW_MAIN)
    c_dir.font = Font(name=FONT_NAME, size=10.5, bold=True, color=C_TEXT_DARK)
    c_dir.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('R1:AG1')
    c_ind = ws['R1']
    c_ind.value = 'INDIRECT LABOUR / SUPPORT'
    c_ind.fill = PatternFill('solid', fgColor=C_YELLOW_SUB)
    c_ind.font = Font(name=FONT_NAME, size=10.5, bold=True, color=C_TEXT_DARK)
    c_ind.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('AH1:AK1')
    c_gt = ws['AH1']
    c_gt.value = 'GRAND TOTAL (PLANT WIDE)'
    c_gt.fill = PatternFill('solid', fgColor=C_YELLOW_ACCENT)
    c_gt.font = Font(name=FONT_NAME, size=10.5, bold=True, color=C_TEXT_DARK)
    c_gt.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[2].height = 22
    sub_cats = [
        (2, 5, 'Operator', C_YELLOW_LIGHT, '78350F'),
        (6, 9, 'Contract Labour', C_YELLOW_LIGHT, '78350F'),
        (10, 13, 'NAPS', C_YELLOW_LIGHT, '78350F'),
        (14, 17, 'Total Direct', C_YELLOW_GOLD, C_TEXT_DARK),
        (18, 21, 'Operator', C_YELLOW_PALE, C_TEXT_DARK),
        (22, 25, 'Contract Labour', C_YELLOW_PALE, C_TEXT_DARK),
        (26, 29, 'NAPS', C_YELLOW_PALE, C_TEXT_DARK),
        (30, 33, 'Total Indirect', C_YELLOW_GOLD, C_TEXT_DARK),
        (34, 37, 'Plant Grand Total', C_YELLOW_TOTAL, C_TEXT_DARK)
    ]
    for s, e, label, bg, fg in sub_cats:
        ws.merge_cells(start_row=2, start_column=s, end_row=2, end_column=e)
        c = ws.cell(row=2, column=s)
        c.value = label
        c.fill = PatternFill('solid', fgColor=bg)
        c.font = Font(name=FONT_NAME, size=9.5, bold=True, color=fg)
        c.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[3].height = 22
    c_dt = ws['A3']
    c_dt.value = 'Date'
    c_dt.fill = PatternFill('solid', fgColor=C_YELLOW_MAIN)
    c_dt.font = Font(name=FONT_NAME, size=9.5, bold=True, color=C_TEXT_DARK)
    c_dt.alignment = Alignment(horizontal='center', vertical='center')
    c_dt.border = thin_border

    for c in range(2, 38):
        mod = (c - 2) % 4
        label = 'Head Count' if mod == 0 else ('CTC' if mod == 1 else ('OT Wages' if mod == 2 else 'Total'))
        cell = ws.cell(row=3, column=c, value=label)
        is_tot = (mod == 3 or c >= 34)
        cell.fill = PatternFill('solid', fgColor=C_YELLOW_ACCENT if is_tot else C_YELLOW_MAIN)
        cell.font = Font(name=FONT_NAME, size=9.5, bold=True, color=C_TEXT_DARK)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    ws.column_dimensions['A'].width = 14
    for c in range(2, 38):
        col_letter = openpyxl.utils.get_column_letter(c)
        ws.column_dimensions[col_letter].width = 11 if (c - 2) % 4 == 0 else 14

    _, num_days = calendar.monthrange(year, month)
    for d in range(1, num_days + 1):
        r = 3 + d
        ws.row_dimensions[r].height = 20
        cell_date = ws.cell(row=r, column=1, value=datetime(year, month, d))
        cell_date.number_format = 'dd-mmm-yyyy'
        cell_date.font = Font(name=FONT_NAME, size=10, color=C_TEXT_SUB)
        cell_date.alignment = Alignment(horizontal='center', vertical='center')
        cell_date.border = thin_border

        row_bg = 'FAFAFA' if d % 2 == 0 else 'FFFFFF'
        for c in range(2, 38):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = Font(name=FONT_NAME, size=10, color='111827')
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.fill = PatternFill('solid', fgColor=row_bg)
            cell.number_format = '#,##0'

def process_all_dates():
    master_file = 'CL CTC Input 2.xlsx'
    if not os.path.exists(master_file): return
    
    wb_master = openpyxl.load_workbook(master_file, data_only=True)
    
    ws_cl = wb_master['Contract']
    rows_cl = list(ws_cl.iter_rows(values_only=True))
    h_cl = [str(c or '').strip().upper() for c in rows_cl[1]]
    idx_emp = h_cl.index('EMP NO')
    cl_map = {}
    for r in rows_cl[2:]:
        if r and r[idx_emp] not in ['', None]:
            cl_map[str(r[idx_emp]).strip().upper()] = {
                'name': r[h_cl.index('NAME')] or '',
                'dept': r[h_cl.index('DEPT')] or '',
                'direct': str(r[h_cl.index('CATEGORY')] or '').strip().upper() == 'DIRECT',
                'dailyCTC': float(r[h_cl.index('DAILY CTC')] or 0),
                'dailyOT': float(r[h_cl.index('DAILY OT')] or 0)
            }

    ws_naps = wb_master['NAPS']
    rows_naps = list(ws_naps.iter_rows(values_only=True))
    h_naps = [str(c or '').strip().upper() for c in rows_naps[0]]
    naps_map = {}
    for r in rows_naps[1:]:
        if r and r[h_naps.index('CODE')] not in ['', None]:
            naps_map[str(r[h_naps.index('CODE')]).strip().upper()] = {
                'name': r[h_naps.index('NAME')] or '',
                'dept': r[h_naps.index('DEPT')] or '',
                'direct': str(r[h_naps.index('DIRECT/INDIRECT')] or '').strip().upper() == 'DIRECT',
                'dailyCTC': float(r[h_naps.index('DAILY CTC')] or 0),
                'dailyOT': 0.0
            }

    ws_op = wb_master['OPERATOR']
    rows_op = list(ws_op.iter_rows(values_only=True))
    h_op = [str(c or '').strip().upper() for c in rows_op[0]]
    op_map = {}
    for r in rows_op[1:]:
        if r and r[h_op.index('EMP CODE')] not in ['', None]:
            op_map[str(r[h_op.index('EMP CODE')]).strip().upper()] = {
                'name': r[h_op.index('EMP NAME')] or '',
                'dept': r[h_op.index('DEPARTMENT')] or '',
                'direct': str(r[h_op.index('DEPARTMENT')] or '').strip().upper() == 'PRODUCTION',
                'dailyCTC': float(r[h_op.index('DAILY CTC')] or 0),
                'dailyOT': float(r[h_op.index('OT')] if 'OT' in h_op else r[h_op.index('DAILY OT')] or 0)
            }

    date_buckets = {}
    for root, dirs, files in os.walk('.'):
        for f in files:
            if not f.endswith('.xlsx') or f.startswith('~$') or 'Output' in root or 'CL CTC' in f or 'Summary' in f or 'Template Update' in f:
                continue
            filepath = os.path.join(root, f)
            try:
                f_date, cat, rows, h_idx = extract_file_info(filepath)
                if not f_date or not cat or h_idx == -1: continue
                h = [str(c or '').strip().upper() for c in rows[h_idx]]
                c_idx = h.index('CODE')
                s_idx = h.index('STATUS')
                o_idx = h.index('OT') if 'OT' in h else -1
                w_idx = h.index('WORKHRS') if 'WORKHRS' in h else (h.index('WORK HRS') if 'WORK HRS' in h else -1)
                
                records = []
                for r in rows[h_idx+1:]:
                    if r and len(r) > c_idx and r[c_idx]:
                        code_str = str(r[c_idx]).strip().upper()
                        if code_str in ['GRANDTOTAL', 'TOTAL RECORDS'] or 'TOTAL' in code_str: continue
                        records.append({
                            'code': code_str,
                            'status': str(r[s_idx] or '').strip().upper() if len(r) > s_idx else '',
                            'otHours': time_str_to_hours(r[o_idx] if o_idx != -1 and len(r) > o_idx else ''),
                            'workHours': time_str_to_hours(r[w_idx] if w_idx != -1 and len(r) > w_idx else '')
                        })
                if f_date not in date_buckets:
                    date_buckets[f_date] = {'CL': [], 'OP': [], 'NAPS': [], 'files': []}
                date_buckets[f_date][cat] = records
            except Exception:
                pass

    sorted_dates = sorted(date_buckets.keys())
    wb_out = openpyxl.Workbook()
    ws_sum = wb_out.active
    ws_sum.title = 'Summary'

    yr = sorted_dates[0].year if sorted_dates else 2026
    mon = sorted_dates[0].month if sorted_dates else 8
    style_summary_sheet(ws_sum, yr, mon)

    emp_stats = {
        'OP': {code: {'workHrs': 0.0, 'daysPresent': 0, 'wopCount': 0, 'otHrs': 0.0, 'wages': 0.0} for code in op_map},
        'CL': {code: {'workHrs': 0.0, 'daysPresent': 0, 'wopCount': 0, 'otHrs': 0.0, 'wages': 0.0} for code in cl_map},
        'NAPS': {code: {'workHrs': 0.0, 'daysPresent': 0, 'wopCount': 0, 'otHrs': 0.0, 'wages': 0.0} for code in naps_map}
    }

    for cur_date in sorted_dates:
        day_num = cur_date.day
        target_sum_row = 3 + day_num
        data = date_buckets[cur_date]
        b = {
            'directOperator': {'headcount': 0, 'ctc': 0.0, 'ot': 0.0},
            'directCL': {'headcount': 0, 'ctc': 0.0, 'ot': 0.0},
            'directNAPS': {'headcount': 0, 'ctc': 0.0, 'ot': 0.0},
            'indirectOperator': {'headcount': 0, 'ctc': 0.0, 'ot': 0.0},
            'indirectCL': {'headcount': 0, 'ctc': 0.0, 'ot': 0.0},
            'indirectNAPS': {'headcount': 0, 'ctc': 0.0, 'ot': 0.0},
        }
        def calc_bucket(records, emp_map, d_key, i_key, stat_dict):
            for rec in records:
                if rec['status'] != 'P' and rec['status'] != 'WOP': continue
                code = rec['code']
                if code in emp_map:
                    info = emp_map[code]
                    bucket = b[d_key] if info['direct'] else b[i_key]
                    bucket['headcount'] += 1
                    bucket['ctc'] += info['dailyCTC']
                    bucket['ot'] += rec['otHours'] * info['dailyOT']
                    if rec['status'] == 'P':
                        stat_dict[code]['daysPresent'] += 1
                    elif rec['status'] == 'WOP':
                        stat_dict[code]['wopCount'] += 1
                    stat_dict[code]['workHrs'] += rec['workHours']
                    stat_dict[code]['otHrs'] += rec['otHours']
                    stat_dict[code]['wages'] += info['dailyCTC'] + (rec['otHours'] * info['dailyOT'])
                    
        calc_bucket(data['OP'], op_map, 'directOperator', 'indirectOperator', emp_stats['OP'])
        calc_bucket(data['CL'], cl_map, 'directCL', 'indirectCL', emp_stats['CL'])
        calc_bucket(data['NAPS'], naps_map, 'directNAPS', 'indirectNAPS', emp_stats['NAPS'])

        dTotOp = b['directOperator']['ctc'] + b['directOperator']['ot']
        dTotCL = b['directCL']['ctc'] + b['directCL']['ot']
        dTotNaps = b['directNAPS']['ctc'] + b['directNAPS']['ot']
        dHC = b['directOperator']['headcount'] + b['directCL']['headcount'] + b['directNAPS']['headcount']
        dCTC = b['directOperator']['ctc'] + b['directCL']['ctc'] + b['directNAPS']['ctc']
        dOT = b['directOperator']['ot'] + b['directCL']['ot'] + b['directNAPS']['ot']
        dTot = dCTC + dOT

        iTotOp = b['indirectOperator']['ctc'] + b['indirectOperator']['ot']
        iTotCL = b['indirectCL']['ctc'] + b['indirectCL']['ot']
        iTotNaps = b['indirectNAPS']['ctc'] + b['indirectNAPS']['ot']
        iHC = b['indirectOperator']['headcount'] + b['indirectCL']['headcount'] + b['indirectNAPS']['headcount']
        iCTC = b['indirectOperator']['ctc'] + b['indirectCL']['ctc'] + b['indirectNAPS']['ctc']
        iOT = b['indirectOperator']['ot'] + b['indirectCL']['ot'] + b['indirectNAPS']['ot']
        iTot = iCTC + iOT

        writes = {
            2: b['directOperator']['headcount'], 3: b['directOperator']['ctc'], 4: b['directOperator']['ot'], 5: dTotOp,
            6: b['directCL']['headcount'],       7: b['directCL']['ctc'],       8: b['directCL']['ot'],       9: dTotCL,
            10: b['directNAPS']['headcount'],    11: b['directNAPS']['ctc'],    12: b['directNAPS']['ot'],    13: dTotNaps,
            14: dHC, 15: dCTC, 16: dOT, 17: dTot,
            18: b['indirectOperator']['headcount'], 19: b['indirectOperator']['ctc'], 20: b['indirectOperator']['ot'], 21: iTotOp,
            22: b['indirectCL']['headcount'],       23: b['indirectCL']['ctc'],       24: b['indirectCL']['ot'],       25: iTotCL,
            26: b['indirectNAPS']['headcount'],     27: b['indirectNAPS']['ctc'],     28: b['indirectNAPS']['ot'],     29: iTotNaps,
            30: iHC, 31: iCTC, 32: iOT, 33: iTot,
            34: dHC + iHC, 35: dCTC + iCTC, 36: dOT + iOT, 37: dTot + iTot
        }
        for col, val in writes.items():
            ws_sum.cell(row=target_sum_row, column=col, value=val)

    # Build Monthly Detail Sheet - Columns A to I (NO Gap, NO Merging)
    def build_monthly_detail_sheet(ws_name, emp_map, stat_dict):
        ws = wb_out.create_sheet(ws_name)
        ws.freeze_panes = 'A2'
        ws.views.sheetView[0].showGridLines = True
        ws.row_dimensions[1].height = 24
        
        headers = {1:'S.No', 2:'Emp Code', 3:'Name', 4:'Department', 5:'Total Work Hrs', 6:'No of days present', 7:'Total WOP Count', 8:'Total OT Hrs', 9:'Wages'}
        for col, h in headers.items():
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = PatternFill('solid', fgColor=C_YELLOW_MAIN)
            cell.font = Font(name=FONT_NAME, size=10, bold=True, color=C_TEXT_DARK)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            
        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 28
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 14
            
        r = 2
        sno = 1
        for code, info in emp_map.items():
            st = stat_dict.get(code, {'workHrs':0.0, 'daysPresent':0, 'wopCount':0, 'otHrs':0.0, 'wages':0.0})
            row_bg = 'FAFAFA' if sno % 2 == 0 else 'FFFFFF'
            
            ws.row_dimensions[r].height = 20
            ws.cell(row=r, column=1, value=sno).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=r, column=2, value=code).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=r, column=3, value=info['name']).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=r, column=4, value=info['dept']).alignment = Alignment(horizontal='center', vertical='center')
            
            c_wh = ws.cell(row=r, column=5, value=round(st['workHrs'], 2) if st['workHrs'] > 0 else 0)
            c_wh.number_format = '#,##0.00'
            c_wh.alignment = Alignment(horizontal='right', vertical='center')
            
            c_dp = ws.cell(row=r, column=6, value=st['daysPresent'])
            c_dp.number_format = '#,##0'
            c_dp.alignment = Alignment(horizontal='center', vertical='center')

            c_wop = ws.cell(row=r, column=7, value=st['wopCount'])
            c_wop.number_format = '#,##0'
            c_wop.alignment = Alignment(horizontal='center', vertical='center')
            
            c_ot = ws.cell(row=r, column=8, value=round(st['otHrs'], 2) if st['otHrs'] > 0 else 0)
            c_ot.number_format = '#,##0.00'
            c_ot.alignment = Alignment(horizontal='right', vertical='center')
            
            c_wg = ws.cell(row=r, column=9, value=round(st['wages'], 2) if st['wages'] > 0 else 0)
            c_wg.number_format = '#,##0'
            c_wg.alignment = Alignment(horizontal='right', vertical='center')

            for c in range(1, 10):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                cell.font = Font(name=FONT_NAME, size=10, color='111827')
                cell.fill = PatternFill('solid', fgColor=row_bg)

            r += 1
            sno += 1

    build_monthly_detail_sheet('ATC', op_map, emp_stats['OP'])
    build_monthly_detail_sheet('CL', cl_map, emp_stats['CL'])
    build_monthly_detail_sheet('NAPS', naps_map, emp_stats['NAPS'])

    os.makedirs('Output', exist_ok=True)
    out_path = os.path.join('Output', 'CTC_Output_August_2026.xlsx')
    wb_out.save(out_path)
    print(f'[SUCCESS] Saved to {out_path}')

if __name__ == '__main__':
    process_all_dates()
